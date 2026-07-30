"""
Regenerates data/<year>.json for rookie-draft-guide from the master Excel workbook
(Final_Rookie_Markers.xlsx). Excel formulas already compute Missing/Markers/Final.Grade/etc;
this script only transcribes those computed values into the site's JSON schema — it does not
reimplement any grading logic. Covers drafted classes only (QB/RB/WR/TE sheets), not the
pre-draft IMPORT staging sheets.

Usage:
    python export_from_workbook.py [--year 2025] [--out-dir data] [--check]

--check compares the freshly exported records against the existing data/<year>.json
(matched by cfr_id) and prints any field where the values differ by more than a small
tolerance, without writing anything.
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl

WORKBOOK_PATH = Path(r"C:\Users\hayes\Documents\Fantasy Database\Final_Rookie_Markers.xlsx")
REPO_DATA_DIR = Path(__file__).resolve().parent.parent / "data"

# Columns A-T are laid out identically across the QB/RB/WR/TE sheets.
COMMON_MAP = {
    "Position": "pos",
    "Player": "name",
    "College": "school",
    "Conference": "conf",
    "CFR_ID": "cfr_id",
    "PFR_ID": "pfr_id",
    "Age": "age",
    "Round": "round",
    "Overall": "overall",
    "Years": "years",
}

POSITION_MAP = {
    "QB": {
        "Career.PA": "careerPA", "Career.PYd": "careerPYd", "Career.PTD": "careerPTD",
        "Career.INT": "careerINT", "Career.CMP%": "careerCMP", "Career.PAPG": "careerPAPG",
        "Career.YPA": "careerYPA", "Career.YPG": "careerYPG", "Career.TDPG": "careerTDPG",
        "Career.INTPG": "careerINTPG", "Career.RuYd": "careerRuYd", "Career.RuYPG": "careerRuYPG",
        "Best.PYd": "bestPYd", "Best.PTD": "bestPTD", "Best.INT": "bestINT",
        "Best.CMP%": "bestCMP", "Best.RuYd": "bestRuYd", "Last.PPG": "lastPPG",
        "Best.PPG": "bestPPG", "Weight": "wt", "Height": "ht", "BMI": "bmi",
        "Career.Acc%": "careerAcc", "Career.BTT": "careerBTTcount", "Career.BTT%": "careerBTT",
        "Career.TWP%": "careerTWP", "Career.Sack%": "careerSackPct", "Best.Acc%": "bestAcc",
        "Best.BTT": "bestBTTcount", "Best.BTT%": "bestBTT", "Best.TWP%": "bestTWP",
        "Best.Sack%": "bestSackPct", "40-Dash": "dash", "WaSS": "wass", "Burst Score": "burst",
    },
    "RB": {
        "YPC": "ypc", "YPTP": "yptp", "Best Receptions": "bestRec",
        "Best Total Yards Y1/Y2": "bestTotalYds", "First PPG": "firstPPG", "Last PPG": "lastPPG",
        "Best PPG": "bestPPG", "Career.BA/A": "careerBA", "Career.MTF/A": "careerMTF",
        "Career.1D/A": "career1D", "Career.YACO/A": "careerYACO", "Career.YPRR": "careerYPRR",
        "Career.TPRR": "careerTPRR", "Best.BA/A": "bestBA", "Best.MTF/A": "bestMTF",
        "Best.1D/A": "best1D", "Best.YACO/A": "bestYACO", "Best.YPRR": "bestYPRR",
        "Best.TPRR": "bestTPRR", "Weight": "wt", "Height": "ht", "BMI": "bmi",
        "40-Dash": "dash", "WaSS": "wass", "Burst Score": "burst",
    },
    "WR": {
        "BOY20": "boy20", "BOY30": "boy30", "BOA20": "boa20", "BOA30": "boa30",
        "First.YPTPA": "firstYPTPA", "Last.YPTPA": "lastYPTPA", "Best.YPTPA": "bestYPTPA",
        "First.MSYards": "firstMSYards", "Best.MSYards": "bestMSYards",
        "Career.YPRR": "careerYPRR", "Career.YPRR.Zone": "careerYPRRzone",
        "Career.TPRR": "careerTPRR", "Career.1DRR": "career1DRR", "Career.YACR": "careerYACR",
        "Career.CTR": "careerCTR", "Career.CR": "careerCR", "Career.MTF": "careerMTF",
        "Career.Slot Rate": "careerSlot", "Best.YPRR": "bestYPRR", "Best.TPRR": "bestTPRR",
        "Best.1DRR": "best1DRR", "Best.YACR": "bestYACR", "High.CTR": "highCTR",
        "High.CR": "highCR", "High.MTF": "highMTF", "High.Slot Rate": "highSlot",
        "Last.PPG": "lastPPG", "Best.PPG": "bestPPG", "Weight": "wt", "Height": "ht",
        "BMI": "bmi", "40-Dash": "dash", "WaSS": "wass", "Burst Score": "burst",
    },
    "TE": {
        "BOY15": "boy15", "BOY20": "boy20", "BOA20": "boa20", "YPR": "careerYPR",
        "YPG": "careerYPG", "First.YPTPA": "firstYPTPA", "Last.YPTPA": "lastYPTPA",
        "Best.YPTPA": "bestYPTPA", "Best.MSYards": "bestMSYards", "Career.YPRR": "careerYPRR",
        "Career.TPRR": "careerTPRR", "Career.1DRR": "career1DRR", "Career.YACR": "careerYACR",
        "Career.CTR": "careerCTR", "Career.CR": "careerCR", "Career.MTF": "careerMTF",
        "Career.Slot Rate": "careerSlot", "Average.PassBlock": "avgPassBlock",
        "Average.RunBlock": "avgRunBlock", "Best.YPRR": "bestYPRR", "Best.TPRR": "bestTPRR",
        "Best.1DRR": "best1DRR", "Best.YACR": "bestYACR", "High.MTF": "highMTF",
        "High.Slot Rate": "highSlot", "Last.PPG": "lastPPG", "Best.PPG": "bestPPG",
        "Weight": "wt", "Height": "ht", "BMI": "bmi", "40-Dash": "dash", "WaSS": "wass",
        "Burst Score": "burst",
    },
}

GRADE_ONE_DECIMAL = {"grade", "finalGrade"}


def round_value(json_key, value):
    if value is None or value == "":
        return None
    if not isinstance(value, (int, float)):
        return value
    if json_key in GRADE_ONE_DECIMAL:
        return round(value, 1)
    if json_key in ("markers", "missing"):
        return int(round(value))
    if isinstance(value, float):
        return round(value, 4)
    return value


def load_adj_markers(wb):
    """Adj.Markers lives only on the Final Rankings sheet, keyed by (year, pos, cfr_id)."""
    ws = wb["Final Rankings"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        year = row[idx["Year"]]
        pos = row[idx["Position"]]
        cfr = row[idx["CFR_ID"]]
        adj = row[idx["Adj.Markers"]]
        if year is not None and cfr:
            out[(year, pos, cfr)] = adj
    return out


def export_position(wb, pos, adj_lookup):
    ws = wb[pos]
    headers = [c.value for c in ws[2]]
    records_by_year = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        raw = {headers[i]: v for i, v in enumerate(row) if headers[i]}
        if not raw.get("Player") or raw.get("Year") is None:
            continue
        p = {}
        p["year"] = raw["Year"]
        for excel_col, json_key in COMMON_MAP.items():
            p[json_key] = round_value(json_key, raw.get(excel_col))
        p["missing"] = round_value("missing", raw.get("Missing"))
        p["markers"] = round_value("markers", raw.get("Markers"))
        p["grade"] = round_value("grade", raw.get("Final.Grade"))
        p["finalGrade"] = p["grade"]
        p["ath"] = round_value("ath", raw.get("Athleticism%"))
        p["prod"] = round_value("prod", raw.get("Production%"))
        p["pff"] = round_value("pff", raw.get("Advanced.Markers"))
        adj = adj_lookup.get((raw["Year"], pos, raw.get("CFR_ID")))
        p["adjMarkers"] = round_value("adjMarkers", adj)
        for excel_col, json_key in POSITION_MAP[pos].items():
            val = round_value(json_key, raw.get(excel_col))
            if val is not None:
                p[json_key] = val
        # drop None-valued common fields to match existing sparse style
        p = {k: v for k, v in p.items() if v is not None}
        records_by_year.setdefault(raw["Year"], []).append(p)
    return records_by_year


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, help="Only export this draft year")
    ap.add_argument("--out-dir", default=str(REPO_DATA_DIR))
    ap.add_argument("--check", action="store_true", help="Diff against existing JSON instead of writing")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    adj_lookup = load_adj_markers(wb)

    all_years = {}
    for pos in ("QB", "RB", "WR", "TE"):
        for year, records in export_position(wb, pos, adj_lookup).items():
            all_years.setdefault(year, []).extend(records)

    years = [args.year] if args.year else sorted(all_years)
    out_dir = Path(args.out_dir)

    for year in years:
        records = all_years.get(year, [])
        if not records:
            print(f"{year}: no rows found in workbook, skipping")
            continue
        if args.check:
            check_against_existing(year, records, out_dir)
        else:
            preserve_existing_adj_markers(records, out_dir / f"{year}.json")
            out_path = out_dir / f"{year}.json"
            out_path.write_text(json.dumps(records, indent=2), encoding="utf-8")
            print(f"{year}: wrote {len(records)} records to {out_path}")


def preserve_existing_adj_markers(records, existing_path):
    """adjMarkers is hand-curated on the Final Rankings tab and only covers a
    subset of prospects. Never overwrite an existing site value with nothing —
    if the workbook lookup came up empty, keep whatever is already published."""
    if not existing_path.exists():
        return
    existing = json.loads(existing_path.read_text(encoding="utf-8"))
    existing_by_key = {(p.get("cfr_id") or f"{p.get('pos')}:{p.get('name')}"): p for p in existing}
    for p in records:
        if "adjMarkers" not in p:
            key = p.get("cfr_id") or f"{p.get('pos')}:{p.get('name')}"
            old = existing_by_key.get(key)
            if old and "adjMarkers" in old:
                p["adjMarkers"] = old["adjMarkers"]


def check_against_existing(year, records, out_dir):
    existing_path = out_dir / f"{year}.json"
    if not existing_path.exists():
        print(f"{year}: no existing file to compare ({len(records)} new records)")
        return
    def key(p):
        return p.get("cfr_id") or f"{p.get('pos')}:{p.get('name')}"

    existing = {key(p): p for p in json.loads(existing_path.read_text(encoding="utf-8"))}
    fresh = {key(p): p for p in records}
    missing_in_fresh = set(existing) - set(fresh)
    missing_in_existing = set(fresh) - set(existing)
    if missing_in_fresh:
        print(f"{year}: {len(missing_in_fresh)} players in existing file but not regenerated: {sorted(missing_in_fresh)[:5]}...")
    if missing_in_existing:
        print(f"{year}: {len(missing_in_existing)} players newly regenerated but not in existing file: {sorted(missing_in_existing)[:5]}...")

    diff_count = 0
    for cfr_id in sorted(set(existing) & set(fresh)):
        old, new = existing[cfr_id], fresh[cfr_id]
        for key in set(old) | set(new):
            ov, nv = old.get(key), new.get(key)
            if isinstance(ov, (int, float)) and isinstance(nv, (int, float)):
                if abs(ov - nv) > max(0.05, abs(ov) * 0.005):
                    print(f"{year} {old.get('name', cfr_id)}: {key} existing={ov} fresh={nv}")
                    diff_count += 1
            elif ov != nv:
                print(f"{year} {old.get('name', cfr_id)}: {key} existing={ov!r} fresh={nv!r}")
                diff_count += 1
    print(f"{year}: {len(existing & fresh if False else set(existing)&set(fresh))} players compared, {diff_count} field diffs")


if __name__ == "__main__":
    sys.exit(main())
